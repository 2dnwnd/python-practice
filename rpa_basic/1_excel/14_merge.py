from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 셀 병합
ws.merge_cells('B2:D2')
ws['B2'].value = 'Mergerd cell'

# 셀 병합해제
ws.unmerge_cells('B2:D2')

wb.save('sample_merge.xlsx')