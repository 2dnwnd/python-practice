from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() # 새로운 sheet 기본이름으로 생성
ws.title = 'MySheet'
ws.sheet_properties.tabColor = 'ff66ff'

ws1 = wb.create_sheet('YourSheet')
ws2 = wb.create_sheet('NewSheet', 2) # 인덱스 2번째로 생성

new_ws = wb['NewSheet'] # 사전 형태로 sheet에 접근가능

print(wb.sheetnames) # 모든 sheet 이름 확인

# sheet 복사
new_ws['A1'] = 'Test'
target = wb.copy_worksheet(new_ws)
target.title = 'Copied sheet'

wb.save('sample.xlsx')