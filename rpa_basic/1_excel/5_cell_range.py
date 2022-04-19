from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기
ws.append(['번호','영어','수학'])
for i in range(1, 11):
    ws.append([i, randint(1,100), randint(1,100)])

# col_B = ws['B'] # 영어점수만 가져오기 B컬럼만 싹다 가져오기
# # print(col_B)
# for cell in col_B:
#     print(cell.value)

# col_range = ws['B:C'] # B부터 C까지 컬럼 가져오기
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

row_title = ws[1]
# for cell in row_title:
#     print(cell.value)

from openpyxl.utils.cell import coordinate_from_string

# row_range = ws[2:ws.max_row]
# for rows in row_range:
#     for cell in rows:
#         # print(cell.value, end=' ')
#         # print(cell.coordinate, end=' ')
#         xy = coordinate_from_string(cell.coordinate)
#         print(xy, end=' ')
#         print(xy[0], end='')
#         print(xy[1], end=' ')
#     print()

# print(tuple(ws.rows))

for row in ws.iter_rows(min_row=1, max_row=5, min_col=2, max_col=3): # 전체 row
    print(row[0].value, row[1].value)

wb.save('sample.xlsx')